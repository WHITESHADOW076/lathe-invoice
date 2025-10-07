import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = "invoices.xlsx"

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Lathe Workshop Invoice System")

        self.jobs = []
        self.amounts = []

        self.setup_excel()
        self.create_widgets()
        self.load_next_bill_no()

    def setup_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            ws.append(["Bill No", "Vehicle No", "Jobs Done", "Amounts", "Total", "Signature"])
            wb.save(EXCEL_FILE)

    def load_next_bill_no(self):
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            max_no = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                bill_no = row[0]
                if bill_no and bill_no.startswith('INV'):
                    try:
                        num = int(bill_no[3:])
                        if num > max_no:
                            max_no = num
                    except:
                        pass
            self.next_bill_no = f"INV{max_no + 1:03d}"
        else:
            self.next_bill_no = "INV001"
        self.bill_no_var.set(self.next_bill_no)

    def create_widgets(self):
        tk.Label(self.root, text="Bill No:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.bill_no_var = tk.StringVar()
        self.bill_no_entry = tk.Entry(self.root, textvariable=self.bill_no_var, state='readonly', width=15)
        self.bill_no_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self.root, text="Vehicle No:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.vehicle_entry = tk.Entry(self.root, width=20)
        self.vehicle_entry.grid(row=1, column=1, padx=5, pady=5)

        # Jobs and Amounts input
        tk.Label(self.root, text="Job Description:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.job_entry = tk.Entry(self.root, width=30)
        self.job_entry.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        tk.Label(self.root, text="Amount (₹):").grid(row=2, column=2, sticky="w", padx=5, pady=5)
        self.amount_entry = tk.Entry(self.root, width=15)
        self.amount_entry.grid(row=2, column=3, sticky="w", padx=5, pady=5)

        self.add_btn = tk.Button(self.root, text="Add Job", command=self.add_job)
        self.add_btn.grid(row=2, column=4, padx=5, pady=5)

        # Jobs List
        tk.Label(self.root, text="Jobs / Repairs Added:").grid(row=3, column=0, sticky="nw", padx=5, pady=5)
        self.jobs_listbox = tk.Listbox(self.root, width=60, height=8)
        self.jobs_listbox.grid(row=3, column=1, columnspan=4, sticky="w", padx=5, pady=5)

        # Total display
        self.total_var = tk.StringVar(value="Total: ₹0.00")
        self.total_label = tk.Label(self.root, textvariable=self.total_var, font=("Arial", 14, "bold"))
        self.total_label.grid(row=4, column=1, sticky="w", padx=5, pady=5)

        # Signature Canvas
        tk.Label(self.root, text="Signature (draw below):").grid(row=5, column=0, sticky="nw", padx=5, pady=5)
        self.sig_canvas = tk.Canvas(self.root, width=400, height=100, bg="white", borderwidth=2, relief="sunken")
        self.sig_canvas.grid(row=5, column=1, columnspan=3, sticky="w", padx=5, pady=5)
        self.sig_canvas.bind('<B1-Motion>', self.draw_signature)
        self.last_x, self.last_y = None, None

        self.clear_sig_btn = tk.Button(self.root, text="Clear Signature", command=self.clear_signature)
        self.clear_sig_btn.grid(row=5, column=4, padx=5, pady=5)

        # Buttons
        self.save_btn = tk.Button(self.root, text="Save Invoice", command=self.save_invoice)
        self.save_btn.grid(row=6, column=1, pady=10)

        self.retrieve_btn = tk.Button(self.root, text="Retrieve Invoice", command=self.retrieve_invoice)
        self.retrieve_btn.grid(row=6, column=2, pady=10)

        self.print_btn = tk.Button(self.root, text="Print Invoice Preview", command=self.print_preview)
        self.print_btn.grid(row=6, column=3, pady=10)

        self.clear_btn = tk.Button(self.root, text="Clear All", command=self.clear_all)
        self.clear_btn.grid(row=6, column=4, pady=10)

    def add_job(self):
        job = self.job_entry.get().strip()
        amount_str = self.amount_entry.get().strip()
        if not job or not amount_str:
            messagebox.showwarning("Input Error", "Please enter both job description and amount.")
            return
        try:
            amount = float(amount_str)
        except ValueError:
            messagebox.showwarning("Input Error", "Amount must be a number.")
            return
        self.jobs.append(job)
        self.amounts.append(amount)
        self.jobs_listbox.insert(tk.END, f"{job} - ₹{amount:.2f}")
        total = sum(self.amounts)
        self.total_var.set(f"Total: ₹{total:.2f}")
        self.job_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)

    def clear_signature(self):
        self.sig_canvas.delete("all")

    def draw_signature(self, event):
        if self.last_x and self.last_y:
            self.sig_canvas.create_line(self.last_x, self.last_y, event.x, event.y, width=2)
        self.last_x = event.x
        self.last_y = event.y
        self.sig_canvas.after(100, self.reset_last_coords)

    def reset_last_coords(self):
        self.last_x, self.last_y = None, None

    def save_invoice(self):
        bill_no = self.bill_no_var.get()
        vehicle_no = self.vehicle_entry.get().strip()
        if not bill_no or not vehicle_no or not self.jobs:
            messagebox.showerror("Error", "Bill Number, Vehicle Number and at least one job are required.")
            return

        total = sum(self.amounts)
        # Extract signature as postscript image
        sig_file = f"signature_{bill_no}.ps"
        self.sig_canvas.postscript(file=sig_file, colormode='color')  # Saves signature as .ps file

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Check duplicated bill number
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == bill_no:
                messagebox.showerror("Error", "Bill number already exists!")
                return

        jobs_str = "; ".join(self.jobs)
        amounts_str = "; ".join(f"{amt:.2f}" for amt in self.amounts)
        ws.append([bill_no, vehicle_no, jobs_str, amounts_str, total, sig_file])
        wb.save(EXCEL_FILE)

        messagebox.showinfo("Success", f"Invoice {bill_no} saved successfully!\nTotal: ₹{total:.2f}")
        self.load_next_bill_no()
        self.clear_all()

    def retrieve_invoice(self):
        bill_no = simpledialog.askstring("Retrieve Invoice", "Enter Bill Number to retrieve:")
        if not bill_no:
            return
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == bill_no:
                self.load_invoice_into_ui(row)
                return
        messagebox.showerror("Not Found", f"No invoice found with Bill No: {bill_no}")

    def load_invoice_into_ui(self, row):
        bill_no, vehicle_no, jobs_str, amounts_str, total, sig_file = row
        self.bill_no_var.set(bill_no)
        self.vehicle_entry.delete(0, tk.END)
        self.vehicle_entry.insert(0, vehicle_no)
        self.jobs_listbox.delete(0, tk.END)
        self.jobs = jobs_str.split("; ")
        self.amounts = list(map(float, amounts_str.split("; ")))
        for job, amt in zip(self.jobs, self.amounts):
            self.jobs_listbox.insert(tk.END, f"{job} - ₹{amt:.2f}")
        self.total_var.set(f"Total: ₹{total:.2f}")

        # Load signature - here we just clear, visualization of .ps requires external view or conversion
        self.clear_signature()
        messagebox.showinfo("Signature", f"Signature saved as file: {sig_file}\nOpen externally to view.")

    def print_preview(self):
        bill_no = self.bill_no_var.get()
        vehicle_no = self.vehicle_entry.get().strip()
        total = sum(self.amounts)
        if not bill_no or not vehicle_no or not self.jobs:
            messagebox.showerror("Error", "Fill in the invoice details first.")
            return
        preview_text = f"Invoice Preview\n\nBill No: {bill_no}\nVehicle No: {vehicle_no}\n\nJobs Done:\n"
        for i, (job, amt) in enumerate(zip(self.jobs, self.amounts), start=1):
            preview_text += f"{i}. {job} - ₹{amt:.2f}\n"
        preview_text += f"\nTotal Amount: ₹{total:.2f}\n\n(Signature Below on Printed Copy)"
        # Show preview in a popup window
        preview_win = tk.Toplevel(self.root)
        preview_win.title(f"Invoice Preview - {bill_no}")
        tk.Message(preview_win, text=preview_text, width=400, font=("Arial", 12)).pack(padx=10, pady=10)

    def clear_all(self):
        self.job_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)
        self.vehicle_entry.delete(0, tk.END)
        self.jobs_listbox.delete(0, tk.END)
        self.jobs.clear()
        self.amounts.clear()
        self.total_var.set("Total: ₹0.00")
        self.clear_signature()
        self.load_next_bill_no()

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()
